import openpyxl
import pandas as pd
import numpy as np
cols_headers = ["SS","Addr","Sel","Name","Vol_Max","Vol_Min","DataSize","EnbBits","BinR","DecR","VolR","BinW","DecW","VolW","Steps"]


def loadxlsm(path):
    data = pd.read_excel(open(path, 'rb'),sheet_name="SPI",usecols="C:M,O:Q,U",skiprows=4,names=cols_headers,dtype=object)
    data = data.dropna(axis=0, subset=['SS'])
    data = data.dropna(axis=0, subset=['Addr'])
    data.reset_index(drop=True, inplace=True)
    #data.index = data.index + 1
    #data.fillna(0)
    #data.SS = data.SS.astype(np.int16)
    #data.Addr = data.Addr.astype(np.int16)
    #data.BitSel = data.BitSel.astype(np.int16)
    #data.DataSize = data.DataSize.astype(np.int16)
    #data.EnbBits = data.EnbBits.astype(np.int16)
    #data.Steps = data.EnbBits.astype(np.int16)
    #data[["DecR","VolR","DecW","VolW"]] = data[["DecR","VolR","DecW","VolW"]].apply(pd.to_numeric)

    return data
    # Sheet --split--> Tuples ----> DataFrame
    #workbook = openpyxl.load_workbook(path)
    #sheet = workbook["SPI"]
    #df = pd.DataFrame(sheet["B6:U205"])
    #print(df)
    #print(type(sheet["B6:U205"][1]))
    #print(sheet["B6:U205"][1])
    #pass
    #print(sheet["B6:U205"][0][0])
    #print(type(sheet["B6:U205"][0][0]))
    #pass
    #for i in range(1, sheet.max_row + 1):
    #    for j in range(1, sheet.max_column + 1):
    #        cell = sheet.cell(row=i, column=j)
    #        if cell.value is not None:
    #            print(cell.value, end=" ")
    #            print(i,j)

if __name__ == '__main__':
    data = loadxlsm("C:\\Users\op\Pictures\Assignment\OSPLL_0723.xlsm")
    print(data)
    print(data.dtypes)
    #for index, row in data.iterrows():
    #    print(row["Steps"])

